from bs4 import BeautifulSoup
import requests
from openpyxl import load_workbook, Workbook
import os

# ---------- FILE SETUP ----------
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
FILE_NAME = os.path.join(BASE_DIR, "BrockClubs.xlsx")
AUTOSAVE_INTERVAL = 10  # save every 10 rows

# ---------- LOAD OR CREATE WORKBOOK ----------
if not os.path.exists(FILE_NAME):
    wb = Workbook()
    sheet = wb.active
    sheet.append(["Club Name", "Email", "Club URL"])
    wb.save(FILE_NAME)

wb = load_workbook(FILE_NAME)
sheet = wb.active

# ---------- LOAD EXISTING CLUBS (DUPLICATE PROTECTION) ----------
existing_clubs = set()
for row in sheet.iter_rows(min_row=2, values_only=True):
    if row[0]:
        existing_clubs.add(row[0].strip().lower())

# ---------- SCRAPE MAIN PAGE ----------
url = "https://www.brockbusu.ca/clubs/browse/"
html_text = requests.get(url).text
page = BeautifulSoup(html_text, "lxml")

club_links = page.find_all("a", class_="msl-gl-link")

rows_added = 0

# ---------- LOOP THROUGH CLUB PAGES ----------
for link in club_links:
    club_url = "https://www.brockbusu.ca" + link["href"]
    soup = BeautifulSoup(requests.get(club_url).text, "lxml")

    # Club name
    name_tag = soup.find("h1", class_="purple")
    club_name = name_tag.text.strip() if name_tag else "N/A"

    # Skip duplicates
    if club_name.lower() in existing_clubs:
        continue

    # Email
    email_tag = soup.find("a", class_="socemail")
    email = (
        email_tag["href"].replace("mailto:", "")
        if email_tag and "mailto:" in email_tag.get("href", "")
        else "N/A"
    )

    print(f"Adding: {club_name} | {email}")

    # Write row
    sheet.append([club_name, email, club_url])
    existing_clubs.add(club_name.lower())
    rows_added += 1

    # ---------- AUTOSAVE ----------
    if rows_added % AUTOSAVE_INTERVAL == 0:
        wb.save(FILE_NAME)
        print("Autosaved...")

# ---------- FINAL SAVE ----------
wb.save(FILE_NAME)
print("Done. Excel file saved.")
